import re
import json
import os
import shutil
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Constants ────────────────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    BASEDIR = Path(os.path.dirname(sys.executable))
else:
    BASEDIR = Path(os.path.dirname(os.path.abspath(__file__)))
CONFIG_FILE = BASEDIR / "config.json"
BACKUP_DIR = BASEDIR / "backup"
EXCEL_DIR = BASEDIR / "excel"

# Color palette
C_BG = "#f0f2f5"
C_CARD = "#ffffff"
C_PRIMARY = "#2563eb"
C_PRIMARY_HOVER = "#1d4ed8"
C_SUCCESS = "#16a34a"
C_DANGER = "#dc2626"
C_DANGER_HOVER = "#b91c1c"
C_WARN = "#f59e0b"
C_TEXT = "#1e293b"
C_SUBTEXT = "#64748b"
C_BORDER = "#e2e8f0"
C_ACCENT_BG = "#eff6ff"

# Zodiac colors (used as tags)
ZODIAC_COLORS = {"鼠": "#3b82f6", "牛": "#3b82f6", "虎": "#3b82f6",
                  "兔": "#ef4444", "龙": "#ef4444", "蛇": "#ef4444",
                  "马": "#ef4444", "羊": "#ef4444",
                  "猴": "#22c55e", "鸡": "#22c55e", "狗": "#22c55e", "猪": "#22c55e"}
WAVE_COLORS = {"蓝": "#3b82f6", "红": "#ef4444", "绿": "#22c55e"}
ZODIAC_SET = {"鼠", "牛", "虎", "兔", "龙", "蛇", "马", "羊", "猴", "鸡", "狗", "猪"}


# ─── Editable Treeview ────────────────────────────────────────────────────────
class EditableTreeview(ttk.Treeview):
    """Treeview that supports double-click to edit cells."""
    def __init__(self, master, on_edit_commit=None, **kwargs):
        super().__init__(master, **kwargs)
        self.on_edit_commit = on_edit_commit
        self._edit_widget = None
        self._edit_col = -1
        self.bind("<Double-1>", self._on_double_click)

    def _on_double_click(self, event):
        if self._edit_widget:
            self._commit_edit()
        region = self.identify_region(event.x, event.y)
        if region != "cell":
            return
        column = self.identify_column(event.x)
        row = self.identify_row(event.y)
        if not row:
            return
        col_idx = int(column.replace("#", "")) - 1
        if col_idx == 0:
            return
        bbox = self.bbox(row, column)
        if not bbox:
            return
        x, y, w, h = bbox
        values = self.item(row, "values")
        current_val = values[col_idx] if col_idx < len(values) else ""
        self._edit_col = col_idx
        self._edit_widget = tk.Entry(self, font=("Microsoft YaHei UI", 10),
                                     relief=tk.FLAT, bd=2,
                                     highlightbackground=C_PRIMARY,
                                     highlightthickness=1)
        self._edit_widget.insert(0, str(current_val))
        self._edit_widget.select_range(0, tk.END)
        self._edit_widget.place(x=x, y=y, width=w, height=h)
        self._edit_widget.bind("<Return>", lambda e: self._commit_edit())
        self._edit_widget.bind("<Escape>", lambda e: self._cancel_edit())
        self._edit_widget.focus_set()

    def _commit_edit(self):
        if not self._edit_widget:
            return
        new_val = self._edit_widget.get().strip()
        if new_val and self._edit_col >= 0:
            selection = self.selection()
            if selection:
                item = selection[0]
                values = list(self.item(item, "values"))
                if self._edit_col < len(values):
                    values[self._edit_col] = new_val
                    self.item(item, values=values)
        self._cancel_edit()
        if self.on_edit_commit:
            self.on_edit_commit()

    def _cancel_edit(self):
        if self._edit_widget:
            self._edit_widget.destroy()
            self._edit_widget = None
            self._edit_col = -1


# ─── Main Application ─────────────────────────────────────────────────────────
class LotteryStatsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("六合彩金额统计系统")
        self.root.geometry("1280x860")
        self.root.minsize(1100, 700)
        self.root.configure(bg=C_BG)

        # ── State ──
        self.config = {}
        self.history = []          # Undo stack
        self.current_data = {}     # {"1": 50, "2": 30, ...}
        self.wb = None             # Current openpyxl workbook
        self._stop_event = threading.Event()
        self._scheduler_thread = None

        # ── Init ──
        EXCEL_DIR.mkdir(exist_ok=True)
        BACKUP_DIR.mkdir(exist_ok=True)
        self.load_config()
        self.load_or_create_excel()
        self.init_data()
        self.build_ui()
        self.refresh_all()
        self.start_midnight_scheduler()

    # ═══════════════════════════════════════════════════════════════════════════
    #  Config
    # ═══════════════════════════════════════════════════════════════════════════
    def load_config(self):
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    self.config = json.load(f)
            except (json.JSONDecodeError, KeyError):
                self.config = {}
        if "生肖映射" not in self.config or len(self.config.get("生肖映射", {})) < 49:
            self.config["生肖映射"] = self._default_zodiac_map()
        self.config.setdefault("界面设置", {"主题色": "#2563eb", "字体大小": 11,
                                            "自动备份": True, "备份保留数量": 30,
                                            "每日重置提醒": True})
        self.config.setdefault("系统状态", {"最后重置日期": ""})
        self.save_config()

    def _default_zodiac_map(self):
        # Real mapping: zodiac → numbers
        zodiac_nums = {
            "鼠": [7, 19, 31, 43],
            "牛": [6, 18, 30, 42],
            "虎": [5, 17, 29, 41],
            "兔": [4, 16, 28, 40],
            "龙": [3, 15, 27, 39],
            "蛇": [2, 14, 26, 38],
            "马": [1, 13, 25, 37, 49],
            "羊": [12, 24, 36, 48],
            "猴": [11, 23, 35, 47],
            "鸡": [10, 22, 34, 46],
            "狗": [9, 21, 33, 45],
            "猪": [8, 20, 32, 44],
        }
        zodiac_wave = {
            "鼠":"蓝","牛":"蓝","虎":"蓝",
            "兔":"红","龙":"红","蛇":"红","马":"红","羊":"红",
            "猴":"绿","鸡":"绿","狗":"绿","猪":"绿",
        }
        m = {}
        for zodiac, nums in zodiac_nums.items():
            wave = zodiac_wave[zodiac]
            for n in nums:
                m[str(n)] = {"生肖": zodiac, "波色": wave}
        # Sort by number
        return {str(k): m[str(k)] for k in sorted(m, key=int)}

    def save_config(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)

    # ═══════════════════════════════════════════════════════════════════════════
    #  Excel
    # ═══════════════════════════════════════════════════════════════════════════
    def get_today_filename(self):
        return f"{datetime.now().strftime('%Y%m%d')}_六合彩金额统计表.xlsx"

    def get_excel_path(self):
        return EXCEL_DIR / self.get_today_filename()

    def load_or_create_excel(self):
        path = self.get_excel_path()
        if path.exists():
            self.wb = openpyxl.load_workbook(path)
            self._read_data_from_excel()
        else:
            self.wb = openpyxl.Workbook()
            self._init_excel_sheets()
            self.save_to_excel()
            # Initialize empty data
            for i in range(1, 50):
                self.current_data[str(i)] = 0

    def _read_data_from_excel(self):
        self.current_data = {}
        for i in range(1, 50):
            self.current_data[str(i)] = 0
        if "统计" in self.wb.sheetnames:
            ws = self.wb["统计"]
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0] is not None and str(row[0]).isdigit():
                    num = str(row[0])
                    raw = str(row[1]) if row[1] is not None else "0"
                    # Strip "元" suffix and any whitespace
                    raw = raw.replace("元", "").strip()
                    try:
                        val = int(raw) if raw.lstrip("-").isdigit() else 0
                    except ValueError:
                        val = 0
                    if 1 <= int(num) <= 49:
                        self.current_data[num] = val

    def _init_excel_sheets(self):
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        # Stats sheet
        ws_stats = self.wb.create_sheet("统计", 0)
        self._style_header(ws_stats, ["数字", "金额（元）"])
        # Zodiac summary sheet
        ws_zodiac = self.wb.create_sheet("生肖统计", 1)
        self._style_header(ws_zodiac, ["生肖", "金额（元）", "波色"])
        # Config sheet (read-only reference)
        ws_cfg = self.wb.create_sheet("生肖配置", 2)
        self._style_header(ws_cfg, ["数字", "生肖", "波色"])
        for num, info in self.config["生肖映射"].items():
            ws_cfg.append([int(num), info["生肖"], info["波色"]])

    def _style_header(self, ws, headers):
        header_font = Font(name="Microsoft YaHei UI", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"))
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def save_to_excel(self):
        if not self.wb:
            return
        path = self.get_excel_path()
        data_font = Font(name="Microsoft YaHei UI", size=10)
        data_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"))

        # ── Stats sheet ──
        if "统计" in self.wb.sheetnames:
            ws = self.wb["统计"]
        else:
            ws = self.wb.create_sheet("统计", 0)
            self._style_header(ws, ["数字", "金额（元）"])
        ws.delete_rows(2, ws.max_row)
        for i in range(1, 50):
            num = str(i)
            amt = self.current_data.get(num, 0)
            r = i + 1
            c1 = ws.cell(row=r, column=1, value=i)
            c2 = ws.cell(row=r, column=2, value=f"{amt}元")
            for c in (c1, c2):
                c.font = data_font
                c.alignment = data_align
                c.border = thin_border
        # Total row
        total_row = 51
        total = sum(self.current_data.values())
        c1 = ws.cell(row=total_row, column=1, value="总计")
        c2 = ws.cell(row=total_row, column=2, value=f"{total}元")
        for c in (c1, c2):
            c.font = Font(name="Microsoft YaHei UI", bold=True, size=11, color="2563EB")
            c.alignment = data_align
            c.border = thin_border

        # ── Zodiac sheet ──
        if "生肖统计" in self.wb.sheetnames:
            ws_z = self.wb["生肖统计"]
        else:
            ws_z = self.wb.create_sheet("生肖统计", 1)
            self._style_header(ws_z, ["生肖", "金额（元）", "波色"])
        ws_z.delete_rows(2, ws_z.max_row)
        zodiac_totals = {}
        for num, amt in self.current_data.items():
            z = self.config["生肖映射"][num]["生肖"]
            c = self.config["生肖映射"][num]["波色"]
            zodiac_totals.setdefault(z, {"amount": 0, "color": c})
            zodiac_totals[z]["amount"] += amt
        row = 2
        for zodiac in ["鼠","牛","虎","兔","龙","蛇","马","羊","猴","鸡","狗","猪"]:
            if zodiac in zodiac_totals:
                d = zodiac_totals[zodiac]
                c1 = ws_z.cell(row=row, column=1, value=zodiac)
                c2 = ws_z.cell(row=row, column=2, value=f"{d['amount']}元")
                c3 = ws_z.cell(row=row, column=3, value=d["color"])
                for c in (c1, c2, c3):
                    c.font = data_font
                    c.alignment = data_align
                    c.border = thin_border
                row += 1
        # Total
        c1 = ws_z.cell(row=row, column=1, value="总计")
        c2 = ws_z.cell(row=row, column=2, value=f"{total}元")
        for c in (c1, c2):
            c.font = Font(name="Microsoft YaHei UI", bold=True, size=11, color="2563EB")
            c.alignment = data_align
            c.border = thin_border

        # ── Config sheet ──
        if "生肖配置" in self.wb.sheetnames:
            ws_c = self.wb["生肖配置"]
        else:
            ws_c = self.wb.create_sheet("生肖配置", 2)
            self._style_header(ws_c, ["数字", "生肖", "波色"])
        ws_c.delete_rows(2, ws_c.max_row)
        for i in range(1, 50):
            num = str(i)
            info = self.config["生肖映射"][num]
            r = i + 1
            ws_c.cell(row=r, column=1, value=i).font = data_font
            ws_c.cell(row=r, column=1).alignment = data_align
            ws_c.cell(row=r, column=1).border = thin_border
            ws_c.cell(row=r, column=2, value=info["生肖"]).font = data_font
            ws_c.cell(row=r, column=2).alignment = data_align
            ws_c.cell(row=r, column=2).border = thin_border
            ws_c.cell(row=r, column=3, value=info["波色"]).font = data_font
            ws_c.cell(row=r, column=3).alignment = data_align
            ws_c.cell(row=r, column=3).border = thin_border

        # Column widths
        for ws_name in ["统计", "生肖统计", "生肖配置"]:
            if ws_name in self.wb.sheetnames:
                ws = self.wb[ws_name]
                for col in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 16

        self.wb.save(path)
        self._update_status(f"已保存 → {self.get_today_filename()}")

    # ═══════════════════════════════════════════════════════════════════════════
    #  Backup & Rollback
    # ═══════════════════════════════════════════════════════════════════════════
    def create_backup(self, tag="手动备份"):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder = BACKUP_DIR / f"{ts}_{tag}"
        folder.mkdir(parents=True, exist_ok=True)
        # Copy config
        shutil.copy2(CONFIG_FILE, folder / "config.json")
        # Copy today's Excel if exists
        excel_path = self.get_excel_path()
        if excel_path.exists():
            shutil.copy2(excel_path, folder / self.get_today_filename())
        # Clean old backups
        self._cleanup_backups()
        self._update_status(f"备份完成 → {ts}_{tag}")
        self.refresh_backup_list()
        return folder

    def _cleanup_backups(self):
        max_backups = self.config["界面设置"].get("备份保留数量", 30)
        backups = sorted(BACKUP_DIR.iterdir(), key=os.path.getmtime, reverse=True)
        for old in backups[max_backups:]:
            if old.is_dir():
                shutil.rmtree(old)

    def restore_backup(self, backup_path):
        if not backup_path.exists():
            messagebox.showerror("错误", "备份文件夹不存在")
            return
        # Restore config
        cfg_src = backup_path / "config.json"
        if cfg_src.exists():
            shutil.copy2(cfg_src, CONFIG_FILE)
            self.load_config()
        # Restore Excel
        for f in backup_path.glob("*.xlsx"):
            dest = EXCEL_DIR / f.name
            shutil.copy2(f, dest)
        # Reload
        self.load_or_create_excel()
        self.refresh_all()
        self._update_status(f"已回滚 → {backup_path.name}")
        messagebox.showinfo("回滚成功", f"已从备份恢复：{backup_path.name}")

    def list_backups(self):
        if not BACKUP_DIR.exists():
            return []
        backups = sorted([d for d in BACKUP_DIR.iterdir() if d.is_dir()],
                         key=os.path.getmtime, reverse=True)
        return backups

    # ═══════════════════════════════════════════════════════════════════════════
    #  Text Parsing
    # ═══════════════════════════════════════════════════════════════════════════
    def parse_text(self):
        text = self.input_box.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("提示", "请输入要解析的文本")
            return
        parsed = self._parse_text_content(text)
        if not parsed:
            messagebox.showinfo("提示", "未识别到有效格式。支持格式：\n  · 1.2.3各5 或 1.2.3各5元\n  · 马.羊各10 或 马.羊各10元\n  · 鼠各20")
            return
        # Save undo snapshot
        snapshot = dict(self.current_data)
        self.history.append(snapshot)
        if len(self.history) > 50:
            self.history.pop(0)
        # Apply
        for num, amt in parsed.items():
            self.current_data[num] = self.current_data.get(num, 0) + amt
        # Persist
        self.save_to_excel()
        if self.config["界面设置"].get("自动备份", True):
            self.create_backup("自动备份")
        self.refresh_all()
        self.undo_btn.config(state=tk.NORMAL)
        self._update_status(f"解析完成 — 识别到 {len(parsed)} 个号码")

    def _parse_text_content(self, text):
        # Normalize separators: Chinese/English commas, whitespace → uniform dots + spaces
        text = text.replace("，", ".").replace(",", ".").replace("、", ".")
        # Convert Chinese number words to digits (longer compounds first)
        cn_nums = [("九十", "90"), ("八十", "80"), ("七十", "70"), ("六十", "60"),
                   ("五十", "50"), ("四十", "40"), ("三十", "30"), ("二十", "20"),
                   ("十", "10"), ("九", "9"), ("八", "8"), ("七", "7"), ("六", "6"),
                   ("五", "5"), ("四", "4"), ("三", "3"), ("二", "2"), ("一", "1")]
        for cn, digit in cn_nums:
            text = text.replace(cn, digit)
        # Remove "元" suffix (but keep "万"/"亿" for amount parsing)
        text = text.replace("元", "")
        result = {}
        groups = re.split(r'\s+', text.strip())
        for group in groups:
            if not group:
                continue
            if "各" in group:
                self._parse_ge_format(group, result)
            else:
                self._parse_zodiac_total_format(group, result)
        return result

    def _parse_ge_format(self, group, result):
        """Parse 'X各Y' format: items separated by dots, each gets the amount."""
        segs = group.split("各")
        for i in range(len(segs) - 1):
            items_str = segs[i].rstrip(".")
            amount_str = segs[i + 1]
            if not items_str:
                continue
            amt_match = re.match(r'(\d+)(万|亿)?', amount_str)
            if not amt_match:
                continue
            amount = int(amt_match.group(1))
            suffix = amt_match.group(2)
            if suffix == "万":
                amount *= 10000
            elif suffix == "亿":
                amount *= 100000000
            for part in items_str.split("."):
                if not part:
                    continue
                # Zodiac name → total amount split across its numbers
                if part in ZODIAC_SET:
                    nums = self._get_numbers_for_zodiac(part)
                    if nums:
                        base = amount // len(nums)
                        rem = amount % len(nums)
                        for idx, num in enumerate(nums):
                            share = base + (1 if idx < rem else 0)
                            result[num] = result.get(num, 0) + share
                else:
                    num = self._resolve_part(part)
                    if num:
                        result[num] = result.get(num, 0) + amount

    def _parse_zodiac_total_format(self, group, result):
        """Parse '平?生肖金额' format: total amount split across zodiac's numbers."""
        zodiac_names = "鼠|牛|虎|兔|龙|蛇|马|羊|猴|鸡|狗|猪"
        pattern = rf'(?:平|特)?({zodiac_names})(\d+)(万|亿)?'
        for match in re.findall(pattern, group):
            zodiac = match[0]
            amount = int(match[1])
            suffix = match[2] if len(match) > 2 else None
            if suffix == "万":
                amount *= 10000
            elif suffix == "亿":
                amount *= 100000000
            nums = self._get_numbers_for_zodiac(zodiac)
            if not nums:
                continue
            # Split amount evenly, remainder goes to first numbers
            base = amount // len(nums)
            rem = amount % len(nums)
            for idx, num in enumerate(nums):
                share = base + (1 if idx < rem else 0)
                result[num] = result.get(num, 0) + share

    def _get_numbers_for_zodiac(self, zodiac):
        """Return sorted list of number strings for a given zodiac."""
        nums = []
        for num, info in self.config["生肖映射"].items():
            if info["生肖"] == zodiac:
                nums.append(num)
        return sorted(nums, key=int)

    def _resolve_part(self, part):
        if part.lstrip("0").isdigit() and 1 <= int(part) <= 49:
            return str(int(part))  # Normalize "08" → "8"
        # Try zodiac name → number lookup
        for num, info in self.config["生肖映射"].items():
            if info["生肖"] == part:
                return num
        # Fuzzy match: strip extra chars, try again
        pure = re.sub(r'[^一-鿿]', '', part)
        if pure:
            for num, info in self.config["生肖映射"].items():
                if info["生肖"] == pure:
                    return num
        return None

    # ═══════════════════════════════════════════════════════════════════════════
    #  Undo
    # ═══════════════════════════════════════════════════════════════════════════
    def undo_last(self):
        if not self.history:
            messagebox.showinfo("提示", "没有可撤销的操作")
            return
        self.current_data = self.history.pop()
        self.save_to_excel()
        self.refresh_all()
        if not self.history:
            self.undo_btn.config(state=tk.DISABLED)
        self._update_status("已撤销上一步操作")

    # ═══════════════════════════════════════════════════════════════════════════
    #  Reset
    # ═══════════════════════════════════════════════════════════════════════════
    def reset_today(self, manual=False):
        if manual and self.config["界面设置"].get("每日重置提醒", True):
            ok = messagebox.askyesno("确认重置", "确定要重置今日所有数据吗？\n\n重置前会自动备份当前数据。")
            if not ok:
                return
        self.create_backup("重置前备份")
        for k in self.current_data:
            self.current_data[k] = 0
        self.history.clear()
        self.config["系统状态"]["最后重置日期"] = datetime.now().strftime("%Y-%m-%d")
        self.save_config()
        self.save_to_excel()
        self.refresh_all()
        self.undo_btn.config(state=tk.DISABLED)
        self._update_status("数据已重置")

    def do_midnight_reset(self):
        today_str = datetime.now().strftime("%Y-%m-%d")
        last = self.config["系统状态"].get("最后重置日期", "")
        if last == today_str:
            return  # Already reset today
        self.reset_today(manual=False)
        # Create new Excel for the new day
        self.load_or_create_excel()
        self.refresh_all()

    # ═══════════════════════════════════════════════════════════════════════════
    #  Midnight Scheduler (background thread)
    # ═══════════════════════════════════════════════════════════════════════════
    def start_midnight_scheduler(self):
        if self._scheduler_thread and self._scheduler_thread.is_alive():
            return
        self._stop_event.clear()
        self._scheduler_thread = threading.Thread(target=self._scheduler_loop, daemon=True)
        self._scheduler_thread.start()

    def _scheduler_loop(self):
        while not self._stop_event.is_set():
            now = datetime.now()
            # Calculate seconds until next midnight + 5 seconds buffer
            next_midnight = (now + timedelta(days=1)).replace(hour=0, minute=0, second=5, microsecond=0)
            wait = (next_midnight - now).total_seconds()
            # Sleep in 30s chunks to allow clean shutdown
            while wait > 0 and not self._stop_event.is_set():
                time.sleep(min(30, wait))
                wait -= 30
            if not self._stop_event.is_set():
                self.root.after(0, self.do_midnight_reset)
                # Update countdown display
                self.root.after(100, self._update_countdown)

    def _update_countdown(self):
        now = datetime.now()
        next_mid = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        remaining = next_mid - now
        h = remaining.seconds // 3600
        m = (remaining.seconds % 3600) // 60
        s = remaining.seconds % 60
        self.countdown_label.config(text=f"⏰ 距下次重置: {h:02d}:{m:02d}:{s:02d}")
        self.root.after(1000, self._update_countdown)

    # ═══════════════════════════════════════════════════════════════════════════
    #  UI Construction
    # ═══════════════════════════════════════════════════════════════════════════
    def build_ui(self):
        fs = self.config["界面设置"].get("字体大小", 11)

        # ── Style ──
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(".", font=("Microsoft YaHei UI", fs), background=C_BG)

        # ── Main container ──
        main = tk.Frame(self.root, bg=C_BG)
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        main.grid_columnconfigure(0, weight=60)
        main.grid_columnconfigure(1, weight=40)
        main.grid_rowconfigure(0, weight=0)
        main.grid_rowconfigure(1, weight=1)
        main.grid_rowconfigure(2, weight=1)

        # ── Header ──
        header = tk.Frame(main, bg=C_BG)
        header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        title = tk.Label(header, text="  六合彩金额统计系统",
                         font=("Microsoft YaHei UI", 22, "bold"),
                         bg=C_BG, fg=C_TEXT)
        title.pack(side=tk.LEFT)
        self.countdown_label = tk.Label(header,
                                        text="⏰ 距下次重置: --:--:--",
                                        font=("Microsoft YaHei UI", 10),
                                        bg=C_BG, fg=C_SUBTEXT)
        self.countdown_label.pack(side=tk.RIGHT, padx=(0, 10))
        self._update_countdown()

        # ── Left Panel: Input + Config ──
        left = tk.Frame(main, bg=C_BG)
        left.grid(row=1, column=0, rowspan=2, sticky="nsew", padx=(0, 6))
        left.grid_columnconfigure(0, weight=1)
        left.grid_rowconfigure(0, weight=0)
        left.grid_rowconfigure(1, weight=1)

        # Input card
        input_card = tk.Frame(left, bg=C_CARD, highlightbackground=C_BORDER,
                              highlightthickness=1, padx=16, pady=14)
        input_card.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        input_card.grid_columnconfigure(0, weight=1)
        tk.Label(input_card, text="📝 数据录入", font=("Microsoft YaHei UI", 13, "bold"),
                 bg=C_CARD, fg=C_TEXT).grid(row=0, column=0, sticky="w", pady=(0, 8))
        tk.Label(input_card,
                 text="支持格式: 1.2.3各5  |  马.羊各10  |  鼠各20  |  1.2.3各5万",
                 font=("Microsoft YaHei UI", 9), bg=C_CARD, fg=C_SUBTEXT).grid(
                     row=1, column=0, sticky="w", pady=(0, 6))
        self.input_box = tk.Text(input_card, height=5,
                                 font=("Microsoft YaHei UI", fs),
                                 bg="#f8fafc", fg=C_TEXT,
                                 relief=tk.FLAT, bd=1,
                                 highlightthickness=1,
                                 highlightbackground=C_BORDER,
                                 wrap=tk.WORD,
                                 padx=8, pady=8)
        self.input_box.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        # Button row
        btn_row = tk.Frame(input_card, bg=C_CARD)
        btn_row.grid(row=3, column=0, sticky="ew")
        self._make_btn(btn_row, "解析文本", self.parse_text, C_PRIMARY, C_PRIMARY_HOVER).pack(
            side=tk.LEFT, padx=(0, 6))
        self.undo_btn = self._make_btn(btn_row, "撤销", self.undo_last, C_DANGER, C_DANGER_HOVER)
        self.undo_btn.pack(side=tk.LEFT, padx=(0, 6))
        self.undo_btn.config(state=tk.DISABLED)
        self._make_btn(btn_row, "备份", lambda: self.create_backup("手动备份"), C_SUCCESS, "#15803d").pack(
            side=tk.LEFT, padx=(0, 6))
        self._make_btn(btn_row, "重置今日", lambda: self.reset_today(manual=True), C_WARN, "#d97706").pack(
            side=tk.LEFT)

        # Config card
        config_card = tk.Frame(left, bg=C_CARD, highlightbackground=C_BORDER,
                               highlightthickness=1, padx=16, pady=14)
        config_card.grid(row=1, column=0, sticky="nsew")
        config_card.grid_columnconfigure(0, weight=1)
        config_card.grid_rowconfigure(0, weight=0)
        config_card.grid_rowconfigure(1, weight=1)
        config_card.grid_rowconfigure(2, weight=0)
        tk.Label(config_card, text="⚙ 生肖配置 (双击编辑)",
                 font=("Microsoft YaHei UI", 13, "bold"),
                 bg=C_CARD, fg=C_TEXT).grid(row=0, column=0, sticky="w", pady=(0, 8))

        # Editable tree
        tree_frame = tk.Frame(config_card, bg=C_CARD)
        tree_frame.grid(row=1, column=0, sticky="nsew")
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        columns = ("数字", "生肖", "波色")
        self.config_tree = EditableTreeview(tree_frame, columns=columns,
                                            show="headings", height=14,
                                            on_edit_commit=self._on_config_edited)
        for ci, col in enumerate(columns):
            self.config_tree.heading(col, text=col)
            w = 60 if col == "数字" else 80
            self.config_tree.column(col, width=w, anchor=tk.CENTER, minwidth=w)
        # Scrollbar
        sb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.config_tree.yview)
        self.config_tree.configure(yscrollcommand=sb.set)
        self.config_tree.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")

        # Style the config tree
        style.configure("Custom.Treeview", rowheight=28, font=("Microsoft YaHei UI", fs))
        style.configure("Custom.Treeview.Heading", font=("Microsoft YaHei UI", fs, "bold"))
        self.config_tree.configure(style="Custom.Treeview")

        # Save config button
        tk.Button(config_card, text="保存配置到文件",
                  command=self._save_config_from_ui,
                  font=("Microsoft YaHei UI", 10),
                  bg=C_SUCCESS, fg="white",
                  relief=tk.FLAT, bd=0,
                  padx=14, pady=6,
                  cursor="hand2",
                  activebackground="#15803d", activeforeground="white").grid(
                      row=2, column=0, pady=(10, 0))

        # ── Right Panel: Stats + Backup ──
        right = tk.Frame(main, bg=C_BG)
        right.grid(row=1, column=1, rowspan=2, sticky="nsew", padx=(6, 0))
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(0, weight=0)
        right.grid_rowconfigure(1, weight=1)

        # Stats card
        stats_card = tk.Frame(right, bg=C_CARD, highlightbackground=C_BORDER,
                              highlightthickness=1, padx=16, pady=14)
        stats_card.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        stats_card.grid_columnconfigure(0, weight=1)
        stats_card.grid_rowconfigure(0, weight=0)
        stats_card.grid_rowconfigure(1, weight=1)
        tk.Label(stats_card, text="📊 实时统计",
                 font=("Microsoft YaHei UI", 13, "bold"),
                 bg=C_CARD, fg=C_TEXT).grid(row=0, column=0, sticky="w", pady=(0, 8))

        # Total display
        self.total_label = tk.Label(stats_card, text="总金额: ¥0 元",
                                    font=("Microsoft YaHei UI", 20, "bold"),
                                    bg=C_CARD, fg=C_PRIMARY)
        self.total_label.grid(row=1, column=0, sticky="w", pady=(0, 10))

        # Stats tree
        stats_cols = ("号码", "生肖", "波色", "金额（元）")
        self.stats_tree = ttk.Treeview(stats_card, columns=stats_cols,
                                       show="headings", height=16)
        self.stats_tree.configure(style="Custom.Treeview")
        for ci, col in enumerate(stats_cols):
            self.stats_tree.heading(col, text=col)
            w = 60 if col == "号码" else 80 if col in ("生肖", "波色") else 110
            self.stats_tree.column(col, width=w, anchor=tk.CENTER, minwidth=w)
        sb2 = ttk.Scrollbar(stats_card, orient=tk.VERTICAL, command=self.stats_tree.yview)
        self.stats_tree.configure(yscrollcommand=sb2.set)
        self.stats_tree.grid(row=2, column=0, sticky="nsew")
        sb2.grid(row=2, column=1, sticky="ns")

        # ── Backup card ──
        backup_card = tk.Frame(right, bg=C_CARD, highlightbackground=C_BORDER,
                               highlightthickness=1, padx=16, pady=14)
        backup_card.grid(row=1, column=0, sticky="nsew")
        backup_card.grid_columnconfigure(0, weight=1)
        backup_card.grid_rowconfigure(0, weight=0)
        backup_card.grid_rowconfigure(1, weight=1)
        backup_card.grid_rowconfigure(2, weight=0)
        tk.Label(backup_card, text="💾 备份管理",
                 font=("Microsoft YaHei UI", 13, "bold"),
                 bg=C_CARD, fg=C_TEXT).grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.backup_listbox = tk.Listbox(backup_card,
                                         font=("Microsoft YaHei UI", 9),
                                         bg="#f8fafc", fg=C_TEXT,
                                         relief=tk.FLAT, bd=1,
                                         highlightthickness=1,
                                         highlightbackground=C_BORDER,
                                         selectbackground=C_PRIMARY,
                                         selectforeground="white",
                                         height=8)
        self.backup_listbox.grid(row=1, column=0, sticky="nsew")
        backup_scroll = ttk.Scrollbar(backup_card, orient=tk.VERTICAL,
                                      command=self.backup_listbox.yview)
        self.backup_listbox.configure(yscrollcommand=backup_scroll.set)
        backup_scroll.grid(row=1, column=1, sticky="ns")

        backup_btn_row = tk.Frame(backup_card, bg=C_CARD)
        backup_btn_row.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        self._make_btn(backup_btn_row, "恢复选中备份", self._restore_selected_backup,
                       C_PRIMARY, C_PRIMARY_HOVER).pack(side=tk.LEFT, padx=(0, 6))
        self._make_btn(backup_btn_row, "删除选中备份", self._delete_selected_backup,
                       C_DANGER, C_DANGER_HOVER).pack(side=tk.LEFT)

        # ── Status bar ──
        status_frame = tk.Frame(self.root, bg=C_BORDER, height=28)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.status_label = tk.Label(status_frame, text="就绪",
                                     font=("Microsoft YaHei UI", 9),
                                     bg="#f8fafc", fg=C_SUBTEXT,
                                     anchor=tk.W, padx=12)
        self.status_label.pack(fill=tk.X)

        # Initial refresh
        self.refresh_backup_list()

    def _make_btn(self, parent, text, command, bg, hover_bg):
        btn = tk.Button(parent, text=text, command=command,
                        font=("Microsoft YaHei UI", 10),
                        bg=bg, fg="white",
                        relief=tk.FLAT, bd=0,
                        padx=14, pady=5,
                        cursor="hand2",
                        activebackground=hover_bg, activeforeground="white")
        return btn

    # ═══════════════════════════════════════════════════════════════════════════
    #  UI Refresh
    # ═══════════════════════════════════════════════════════════════════════════
    def refresh_all(self):
        self.refresh_config_tree()
        self.refresh_stats()
        self.refresh_backup_list()

    def refresh_config_tree(self):
        for item in self.config_tree.get_children():
            self.config_tree.delete(item)
        for num, info in self.config["生肖映射"].items():
            self.config_tree.insert("", tk.END,
                                    values=(num, info["生肖"], info["波色"]))

    def refresh_stats(self):
        for item in self.stats_tree.get_children():
            self.stats_tree.delete(item)
        total = 0
        zodiac_totals = {}
        for i in range(1, 50):
            num = str(i)
            amt = self.current_data.get(num, 0)
            total += amt
            zi = self.config["生肖映射"][num]
            zodiac_totals.setdefault(zi["生肖"], {"amount": 0, "color": zi["波色"]})
            zodiac_totals[zi["生肖"]]["amount"] += amt
            tag = f"wave_{zi['波色']}"
            self.stats_tree.insert("", tk.END,
                                   values=(i, zi["生肖"], zi["波色"], f"{amt}元"),
                                   tags=(tag,))
        # Color tags
        self.stats_tree.tag_configure("wave_蓝", foreground="#3b82f6")
        self.stats_tree.tag_configure("wave_红", foreground="#ef4444")
        self.stats_tree.tag_configure("wave_绿", foreground="#22c55e")
        self.total_label.config(text=f"总金额: ¥{total} 元")

    def refresh_backup_list(self):
        self.backup_listbox.delete(0, tk.END)
        for b in self.list_backups():
            ts_raw = b.name[:15]
            try:
                dt = datetime.strptime(ts_raw[:15], "%Y%m%d_%H%M%S")
                display = dt.strftime("%Y-%m-%d %H:%M:%S") + b.name[15:]
            except ValueError:
                display = b.name
            self.backup_listbox.insert(tk.END, display)

    def _update_status(self, msg):
        if hasattr(self, "status_label") and self.status_label:
            self.status_label.config(text=msg)

    # ═══════════════════════════════════════════════════════════════════════════
    #  UI Callbacks
    # ═══════════════════════════════════════════════════════════════════════════
    def _on_config_edited(self):
        self._save_config_from_ui()

    def _save_config_from_ui(self):
        for item in self.config_tree.get_children():
            vals = self.config_tree.item(item, "values")
            if len(vals) >= 3:
                num = str(vals[0])
                self.config["生肖映射"][num] = {"生肖": vals[1], "波色": vals[2]}
        self.save_config()
        # Rebuild Excel config sheet
        if self.wb:
            ws = self.wb["生肖配置"] if "生肖配置" in self.wb.sheetnames else self.wb.create_sheet("生肖配置")
            ws.delete_rows(2, ws.max_row)
            for num, info in self.config["生肖映射"].items():
                ws.append([int(num), info["生肖"], info["波色"]])
            self.wb.save(self.get_excel_path())
        self.refresh_stats()
        self._update_status("配置已保存")

    def _restore_selected_backup(self):
        sel = self.backup_listbox.curselection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个备份")
            return
        backups = self.list_backups()
        idx = sel[0]
        if idx < len(backups):
            ok = messagebox.askyesno("确认回滚",
                                     f"确定要恢复到备份：{backups[idx].name}？\n\n当前数据将被覆盖。")
            if ok:
                self.restore_backup(backups[idx])

    def _delete_selected_backup(self):
        sel = self.backup_listbox.curselection()
        if not sel:
            messagebox.showwarning("提示", "请先选择一个备份")
            return
        backups = self.list_backups()
        idx = sel[0]
        if idx < len(backups):
            ok = messagebox.askyesno("确认删除", f"确定要删除备份：{backups[idx].name}？")
            if ok:
                shutil.rmtree(backups[idx])
                self.refresh_backup_list()
                self._update_status("备份已删除")

    def init_data(self):
        for i in range(1, 50):
            num = str(i)
            if num not in self.current_data:
                self.current_data[num] = 0

    # ═══════════════════════════════════════════════════════════════════════════
    #  Lifecycle
    # ═══════════════════════════════════════════════════════════════════════════
    def on_closing(self):
        self._stop_event.set()
        if self.wb:
            try:
                self.save_to_excel()
            except Exception:
                pass
        self.root.destroy()


# ─── Entry Point ──────────────────────────────────────────────────────────────
def main():
    root = tk.Tk()
    app = LotteryStatsApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
